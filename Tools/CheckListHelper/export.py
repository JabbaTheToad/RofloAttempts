import os
from datetime import datetime
from config import get_exports_dir

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.fonts import addMapping
    from reportlab.lib.enums import TA_LEFT, TA_CENTER
    from reportlab.platypus import KeepTogether
    import os

    PDF_AVAILABLE = True

    # Регистрируем шрифт с поддержкой кириллицы
    # Пробуем разные варианты системных шрифтов
    font_paths = [
        ("DejaVuSans", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),  # Linux
        ("DejaVuSans", "C:/Windows/Fonts/DejaVuSans.ttf"),  # Windows (если установлен)
        ("Arial", "C:/Windows/Fonts/arial.ttf"),  # Windows Arial
        ("TimesNewRoman", "C:/Windows/Fonts/times.ttf"),  # Windows Times New Roman
        ("Verdana", "C:/Windows/Fonts/verdana.ttf"),  # Windows Verdana
    ]

    FONT_REGISTERED = False
    for font_name, font_path in font_paths:
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont(font_name, font_path))
                # Создаем отображение для жирного и курсивного вариантов
                addMapping(font_name, 0, 0, font_name)  # normal
                addMapping(font_name, 1, 0, font_name)  # bold
                addMapping(font_name, 0, 1, font_name)  # italic
                addMapping(font_name, 1, 1, font_name)  # bold italic
                FONT_REGISTERED = True
                DEFAULT_FONT = font_name
                break
            except:
                continue

    if not FONT_REGISTERED:
        # Если не нашли системный шрифт, используем стандартный (может не работать с кириллицей)
        DEFAULT_FONT = 'Helvetica'
        print("Предупреждение: Не найден шрифт с поддержкой кириллицы. Русский текст может отображаться некорректно.")

except ImportError:
    PDF_AVAILABLE = False
    DEFAULT_FONT = 'Helvetica'


class ExportManager:
    """Менеджер экспорта данных"""

    def __init__(self):
        self.exports_dir = get_exports_dir()

    def export_to_excel(self, data):
        """Экспортирует данные в Excel"""
        if not EXCEL_AVAILABLE:
            return False, "Библиотека openpyxl не установлена. Установите: pip install openpyxl"

        try:
            # Создаем имя файла
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"checklist_export_{timestamp}.xlsx"
            filepath = os.path.join(self.exports_dir, filename)

            # Создаем workbook
            wb = openpyxl.Workbook()

            # Удаляем стандартный лист
            wb.remove(wb.active)

            # Стили
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            done_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            bug_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Информация о проекте
            info_ws = wb.create_sheet("Информация")
            info_data = [
                ["Параметр", "Значение"],
                ["Проект", data.get("project_name", "—")],
                ["Версия", data.get("project_version", "—")],
                ["Дата экспорта", data.get("timestamp", "—")],
                ["Тип экспорта", {
                    "project_common": "Общие чек-листы проекта",
                    "object": "Отдельный объект",
                    "full_project": "Весь проект"
                }.get(data.get("type"), "—")]
            ]

            for i, row in enumerate(info_data):
                for j, value in enumerate(row):
                    cell = info_ws.cell(row=i + 1, column=j + 1, value=value)
                    if i == 0:  # Заголовок
                        cell.font = header_font
                        cell.fill = header_fill
                    cell.border = border

            info_ws.column_dimensions['A'].width = 20
            info_ws.column_dimensions['B'].width = 40

            # Создаем листы для каждой секции
            for section in data.get("sections", []):
                ws = wb.create_sheet(section["name"][:30])  # Ограничиваем длину имени листа

                row_idx = 1

                # Заголовок секции
                ws.merge_cells(f'A{row_idx}:D{row_idx}')
                title_cell = ws.cell(row=row_idx, column=1, value=section["name"])
                title_cell.font = Font(bold=True, size=14)
                title_cell.alignment = Alignment(horizontal='center')
                row_idx += 2

                for tab in section.get("tabs", []):
                    # Заголовок вкладки
                    ws.cell(row=row_idx, column=1, value=tab["name"]).font = Font(bold=True, size=12)
                    row_idx += 1

                    # Заголовки таблицы
                    headers = ["Пункт", "Статус", "Комментарий"]
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row_idx, column=col, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.border = border
                    row_idx += 1

                    # Данные
                    for item in tab.get("items", []):
                        # Пункт
                        ws.cell(row=row_idx, column=1, value=item["name"]).border = border

                        # Статус
                        status_cell = ws.cell(row=row_idx, column=2, value=item["status_text"])
                        status_cell.border = border
                        if item["status"] == 1:
                            status_cell.fill = done_fill
                        elif item["status"] == 2:
                            status_cell.fill = bug_fill

                        # Комментарий
                        comment_cell = ws.cell(row=row_idx, column=3, value=item["comment"])
                        comment_cell.border = border
                        comment_cell.alignment = Alignment(wrap_text=True)  # Включаем перенос текста

                        row_idx += 1

                    row_idx += 1

                # Настройка ширины колонок
                ws.column_dimensions['A'].width = 50
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 60  # Увеличиваем ширину колонки для комментариев

                # Включаем автоматический перенос для всех ячеек
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.column == 3:  # Колонка комментариев
                            cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Сохраняем файл
            wb.save(filepath)
            return True, filepath

        except Exception as e:
            return False, str(e)

    def export_to_pdf(self, data):
        """Экспортирует данные в PDF с поддержкой русского языка и длинных комментариев"""
        if not PDF_AVAILABLE:
            return False, "Библиотека reportlab не установлена. Установите: pip install reportlab"

        try:
            # Создаем имя файла
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"checklist_export_{timestamp}.pdf"
            filepath = os.path.join(self.exports_dir, filename)

            # Создаем PDF документ с увеличенными отступами
            doc = SimpleDocTemplate(filepath, pagesize=A4,
                                    leftMargin=40, rightMargin=40,
                                    topMargin=40, bottomMargin=40)
            styles = getSampleStyleSheet()
            elements = []

            # Создаем стили с поддержкой русского шрифта
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontName=DEFAULT_FONT,
                fontSize=16,
                spaceAfter=30,
                alignment=TA_CENTER,
                encoding='utf-8'
            )

            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontName=DEFAULT_FONT,
                fontSize=10,
                leading=14,  # Межстрочный интервал
                alignment=TA_LEFT,
                encoding='utf-8',
                wordWrap='CJK'  # Включаем перенос слов для кириллицы
            )

            heading2_style = ParagraphStyle(
                'CustomHeading2',
                parent=styles['Heading2'],
                fontName=DEFAULT_FONT,
                fontSize=14,
                spaceBefore=20,
                spaceAfter=10,
                encoding='utf-8'
            )

            heading3_style = ParagraphStyle(
                'CustomHeading3',
                parent=styles['Heading3'],
                fontName=DEFAULT_FONT,
                fontSize=12,
                spaceBefore=10,
                spaceAfter=5,
                encoding='utf-8'
            )

            # Стиль для текста в ячейках таблицы
            cell_style = ParagraphStyle(
                'CellStyle',
                parent=styles['Normal'],
                fontName=DEFAULT_FONT,
                fontSize=9,
                leading=12,
                alignment=TA_LEFT,
                encoding='utf-8',
                wordWrap='CJK'  # Включаем перенос слов
            )

            # Заголовок
            elements.append(Paragraph(f"Отчет по тестированию", title_style))
            elements.append(Spacer(1, 0.2 * inch))

            # Информация о проекте
            elements.append(Paragraph(f"<b>Проект:</b> {data.get('project_name', '—')}", normal_style))
            elements.append(Paragraph(f"<b>Версия:</b> {data.get('project_version', '—')}", normal_style))
            elements.append(Paragraph(f"<b>Дата экспорта:</b> {data.get('timestamp', '—')}", normal_style))
            elements.append(Spacer(1, 0.2 * inch))

            # Данные секций
            for section in data.get("sections", []):
                # Заголовок секции
                elements.append(Paragraph(section["name"], heading2_style))

                for tab in section.get("tabs", []):
                    # Заголовок вкладки
                    elements.append(Paragraph(tab["name"], heading3_style))

                    # Подготавливаем данные для таблицы
                    table_data = []

                    # Заголовок таблицы
                    table_data.append([
                        Paragraph("<b>Пункт</b>", cell_style),
                        Paragraph("<b>Статус</b>", cell_style),
                        Paragraph("<b>Комментарий</b>", cell_style)
                    ])

                    # Данные
                    for item in tab.get("items", []):
                        # Обрабатываем длинные комментарии
                        comment_text = item["comment"] if item["comment"] else ""

                        # Разбиваем длинный комментарий на строки с переносом
                        table_data.append([
                            Paragraph(item["name"], cell_style),
                            Paragraph(item["status_text"], cell_style),
                            Paragraph(comment_text, cell_style)
                        ])

                    # Создаем таблицу с оптимальной шириной колонок
                    col_widths = [3.2 * inch, 0.8 * inch, 2.5 * inch]  # Увеличиваем колонку комментария
                    table = Table(table_data, colWidths=col_widths, repeatRows=1)

                    # Стили таблицы
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Выравнивание по верхнему краю
                        ('FONTNAME', (0, 0), (-1, -1), DEFAULT_FONT),
                        ('FONTSIZE', (0, 0), (-1, -1), 9),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('WORDWRAP', (0, 0), (-1, -1), True),  # Включаем перенос слов
                        ('LEFTPADDING', (0, 0), (-1, -1), 6),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                        ('TOPPADDING', (0, 0), (-1, -1), 6),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ]))

                    # Цвета для статусов
                    for i, row in enumerate(table_data[1:], start=1):
                        status_text = row[1].text  # Получаем текст статуса из Paragraph
                        if "Done" in status_text:
                            table.setStyle(TableStyle([
                                ('BACKGROUND', (1, i), (1, i), colors.lightgreen)
                            ]))
                        elif "BUG" in status_text:
                            table.setStyle(TableStyle([
                                ('BACKGROUND', (1, i), (1, i), colors.lightcoral)
                            ]))

                    elements.append(table)
                    elements.append(Spacer(1, 0.1 * inch))

                elements.append(Spacer(1, 0.2 * inch))

            # Строим PDF
            doc.build(elements)
            return True, filepath

        except Exception as e:
            return False, str(e)